import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from copy import copy
import os
import math
import numpy as np
import scipy.stats as stats

def copy_cell(src_cell, dst_cell):
    dst_cell.value = src_cell.value
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.border = copy(src_cell.border)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy(src_cell.protection)

def calculate_group_stats(data_rows, col_idx):
    values = []
    for row in data_rows:
        val = row[col_idx - 1]
        if val is not None:
            try:
                values.append(float(val))
            except ValueError:
                pass
    if not values:
        return {"mean": 0.0, "sem": 0.0, "values": []}
    
    mean_val = np.mean(values)
    if len(values) > 1:
        sem_val = stats.sem(values)
    else:
        sem_val = 0.0
    return {"mean": mean_val, "sem": sem_val, "values": values}

def get_p_value(group_a_vals, group_b_vals):
    if len(group_a_vals) < 2 or len(group_b_vals) < 2:
        return float('nan')
    t_stat, p_val = stats.ttest_ind(group_a_vals, group_b_vals, equal_var=True)
    return p_val

def format_p_value(p_val):
    if math.isnan(p_val):
        return "N/A"
    if p_val >= 0.0001:
        p_str = f"{p_val:.4f}"
    else:
        p_str = "<0.0001"
    
    if p_val < 0.0001:
        p_str += " ****"
    elif p_val < 0.001:
        p_str += " ***"
    elif p_val < 0.01:
        p_str += " **"
    elif p_val < 0.05:
        p_str += " *"
    else:
        p_str += " (ns)"
    return p_str

def apply_summary_styles(sheet, start_row, max_col):
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    header_fill = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")
    header_font = Font(name="Arial", size=10, bold=True)
    section_font = Font(name="Arial", size=11, bold=True, color="003366")
    
    sheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=max_col)
    title_cell = sheet.cell(row=start_row, column=1)
    title_cell.font = section_font
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[start_row].height = 25
    
    sheet.row_dimensions[start_row + 1].height = 20
    for col in range(1, max_col + 1):
        cell = sheet.cell(row=start_row + 1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

def stack_sheets():
    cleaned_dir = "/Users/macboook/PROJECTS/Final year project work/bay-leaf-clove-toxicity-study/cook/cleaned"
    clove_fp = os.path.join(cleaned_dir, "PHILIP OXIDATIVE STRESS MARKERS RESULTS (2026).xlsx")
    bay_fp = os.path.join(cleaned_dir, "PHILIP OXIDATIVE STRESS MARKERS RESULTS 2 (2026).xlsx")
    output_fp = "/Users/macboook/PROJECTS/Final year project work/bay-leaf-clove-toxicity-study/cook/OXIDATIVE_STRESS_MARKERS_STACKED.xlsx"
    
    # 1. Load source workbooks
    wb_clove = openpyxl.load_workbook(clove_fp, data_only=True)
    sheet_clove = wb_clove.active
    
    wb_bay = openpyxl.load_workbook(bay_fp, data_only=True)
    sheet_bay = wb_bay.active
    
    # 2. Create target workbook and sheet
    wb_new = openpyxl.Workbook()
    sheet_new = wb_new.active
    sheet_new.title = "Consolidated_Data"
    
    # Copy headers (rows 1 and 2) from Clove sheet
    for r in [1, 2]:
        sheet_new.row_dimensions[r].height = sheet_clove.row_dimensions[r].height
        for c in range(1, 9):
            copy_cell(sheet_clove.cell(row=r, column=c), sheet_new.cell(row=r, column=c))
            
    # Set column widths based on Clove sheet
    for c in range(1, 9):
        col_letter = openpyxl.utils.get_column_letter(c)
        if col_letter in sheet_clove.column_dimensions:
            sheet_new.column_dimensions[col_letter].width = sheet_clove.column_dimensions[col_letter].width
            
    # 3. Read Clove data rows (Row 3 to 31) and append
    clove_rows = []
    dest_row = 3
    for src_row in range(3, 32):
        sheet_new.row_dimensions[dest_row].height = sheet_clove.row_dimensions[src_row].height
        for c in range(1, 9):
            copy_cell(sheet_clove.cell(row=src_row, column=c), sheet_new.cell(row=dest_row, column=c))
        
        # Save values for statistics
        clove_rows.append([sheet_clove.cell(row=src_row, column=c).value for c in range(1, 9)])
        dest_row += 1
        
    # 4. Read Bay Leaf data rows (Row 3 to 28) and append
    bay_rows = []
    active_prefix = None
    for src_row in range(3, 29):
        sheet_new.row_dimensions[dest_row].height = sheet_bay.row_dimensions[src_row].height
        
        # Read the row values
        row_vals = [sheet_bay.cell(row=src_row, column=c).value for c in range(1, 9)]
        
        # Parse group name for clean labeling
        row_id_val = str(row_vals[0]).strip()
        has_letters = any(char.isalpha() for char in row_id_val)
        if has_letters:
            prefix = ""
            for char in row_id_val:
                if char.isalpha():
                    prefix += char
                elif char.isspace() or char.isdigit():
                    break
            active_prefix = prefix
            clean_id = row_id_val
        else:
            # Numeric ID, append prefix
            clean_id = f"{active_prefix} {row_id_val}"
            
        # Copy values and styling
        for c in range(1, 9):
            copy_cell(sheet_bay.cell(row=src_row, column=c), sheet_new.cell(row=dest_row, column=c))
            
        # Override the ID column with clean fully-written label
        sheet_new.cell(row=dest_row, column=1, value=clean_id)
        
        # Save values for statistics
        bay_rows.append((active_prefix, [clean_id] + row_vals[1:]))
        dest_row += 1
        
    print(f"Stacked raw data. Total data rows: {dest_row - 3}")
    
    # 5. Extract groups for statistical analysis
    # Groups from Clove: CL (CL1-15), MC (MC1-5), NC (NC1-5), IB (IB1,3,4,5)
    clove_cl = [row for row in clove_rows if str(row[0]).startswith("CL")]
    clove_mc = [row for row in clove_rows if str(row[0]).startswith("MC")]
    clove_nc = [row for row in clove_rows if str(row[0]).startswith("NC")]
    clove_ib = [row for row in clove_rows if str(row[0]).startswith("IB")]
    
    # Groups from Bay Leaf: BL (BL1-15), IBU (IBU1,2,4), MG (MG1-3), NG (NG1-3), CL (CL16-17)
    bay_bl = [row[1] for row in bay_rows if row[0] == "BL"]
    bay_ibu = [row[1] for row in bay_rows if row[0] == "IBU"]
    bay_mg = [row[1] for row in bay_rows if row[0] == "MG"]
    bay_ng = [row[1] for row in bay_rows if row[0] == "NG"]
    bay_cl = [row[1] for row in bay_rows if row[0] == "CL"]
    
    # Pooled Clove group (CL1-15 + CL16-17)
    pooled_cl = clove_cl + bay_cl
    
    print(f"Clove run groups: CL={len(clove_cl)}, MC={len(clove_mc)}, NC={len(clove_nc)}, IB={len(clove_ib)}")
    print(f"Bay Leaf run groups: BL={len(bay_bl)}, IBU={len(bay_ibu)}, MG={len(bay_mg)}, NG={len(bay_ng)}, CL={len(bay_cl)}")
    print(f"Pooled Clove group size: {len(pooled_cl)}")
    
    # Compute stats for all groups
    all_stats = {}
    for col in range(2, 9):
        nc_s = calculate_group_stats(clove_nc, col)
        mc_s = calculate_group_stats(clove_mc, col)
        cl_s = calculate_group_stats(pooled_cl, col)
        ib_s = calculate_group_stats(clove_ib, col)
        
        ng_s = calculate_group_stats(bay_ng, col)
        mg_s = calculate_group_stats(bay_mg, col)
        bl_s = calculate_group_stats(bay_bl, col)
        ibu_s = calculate_group_stats(bay_ibu, col)
        
        # p-values
        p_mc_nc = get_p_value(mc_s["values"], nc_s["values"])
        p_cl_mc = get_p_value(cl_s["values"], mc_s["values"])
        p_ib_mc = get_p_value(ib_s["values"], mc_s["values"])
        
        p_mg_ng = get_p_value(mg_s["values"], ng_s["values"])
        p_bl_mg = get_p_value(bl_s["values"], mg_s["values"])
        p_ibu_mg = get_p_value(ibu_s["values"], mg_s["values"])
        
        all_stats[col] = {
            "nc": nc_s, "mc": mc_s, "cl": cl_s, "ib": ib_s,
            "ng": ng_s, "mg": mg_s, "bl": bl_s, "ibu": ibu_s,
            "p_mc_nc": p_mc_nc, "p_cl_mc": p_cl_mc, "p_ib_mc": p_ib_mc,
            "p_mg_ng": p_mg_ng, "p_bl_mg": p_bl_mg, "p_ibu_mg": p_ibu_mg
        }
        
    # 6. Append Summary Table at the bottom
    start_row = dest_row + 2
    sheet_new.cell(row=start_row, column=1, value="STATISTICAL ANALYSIS SUMMARY")
    
    # Headers
    headers = ["Group / Parameter"] + [sheet_clove.cell(row=1, column=c).value for c in range(2, 9)]
    for idx, h in enumerate(headers, 1):
        sheet_new.cell(row=start_row + 1, column=idx, value=h)
        
    apply_summary_styles(sheet_new, start_row, 8)
    
    # Means rows
    groups_to_write = [
        ("NC (Normal Control - Clove Run)", "nc"),
        ("MC (Model Control - Clove Run)", "mc"),
        ("CL (Clove Extract - Pooled CL 1-17)", "cl"),
        ("IB (Ibuprofen - Clove Run)", "ib"),
        ("NG (Normal Control - Bay Leaf Run)", "ng"),
        ("MG (Model Control - Bay Leaf Run)", "mg"),
        ("BL (Bay Leaf Extract)", "bl"),
        ("IBU (Ibuprofen - Bay Leaf Run)", "ibu")
    ]
    
    border_cell = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    font_bold = Font(name="Arial", size=9, bold=True)
    font_regular = Font(name="Arial", size=9, bold=False)
    
    curr_row = start_row + 2
    for label, key in groups_to_write:
        sheet_new.cell(row=curr_row, column=1, value=label).font = font_bold
        sheet_new.cell(row=curr_row, column=1).alignment = Alignment(horizontal="left")
        sheet_new.cell(row=curr_row, column=1).border = border_cell
        sheet_new.row_dimensions[curr_row].height = 18
        
        for col in range(2, 9):
            stat = all_stats[col][key]
            mean_sem_str = f"{stat['mean']:.4f} ± {stat['sem']:.4f}"
            cell = sheet_new.cell(row=curr_row, column=col, value=mean_sem_str)
            cell.font = font_regular
            cell.alignment = Alignment(horizontal="center")
            cell.border = border_cell
        curr_row += 1
        
    # p-values rows
    p_values_to_write = [
        ("p-value (MC vs NC)", "p_mc_nc"),
        ("p-value (CL vs MC)", "p_cl_mc"),
        ("p-value (IB vs MC)", "p_ib_mc"),
        ("p-value (MG vs NG)", "p_mg_ng"),
        ("p-value (BL vs MG)", "p_bl_mg"),
        ("p-value (IBU vs MG)", "p_ibu_mg")
    ]
    
    for label, key in p_values_to_write:
        sheet_new.cell(row=curr_row, column=1, value=label).font = font_bold
        sheet_new.cell(row=curr_row, column=1).alignment = Alignment(horizontal="left")
        sheet_new.cell(row=curr_row, column=1).border = border_cell
        sheet_new.row_dimensions[curr_row].height = 18
        
        for col in range(2, 9):
            p_val = all_stats[col][key]
            p_str = format_p_value(p_val)
            cell = sheet_new.cell(row=curr_row, column=col, value=p_str)
            cell.font = font_regular
            cell.alignment = Alignment(horizontal="center")
            cell.border = border_cell
        curr_row += 1
        
    # Save target workbook
    wb_new.save(output_fp)
    print(f"Stacked workbook saved successfully to: {output_fp}")

if __name__ == "__main__":
    stack_sheets()
