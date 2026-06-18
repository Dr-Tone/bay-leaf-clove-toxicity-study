import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from copy import copy
import os
import math
import numpy as np
import scipy.stats as stats

def copy_cell(src_cell, dst_cell):
    dst_cell.value = src_cell.value
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.border = copy(src_cell.border)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy(src_cell.protection)

def calculate_group_stats(data_rows, col_idx):
    # Extract numeric values for the given column index (1-based)
    values = []
    for row in data_rows:
        val = row[col_idx - 1]
        if val is not None:
            try:
                values.append(float(val))
            except ValueError:
                pass
    
    if not values:
        return {"mean": 0.0, "sem": 0.0, "values": []}
    
    mean_val = np.mean(values)
    
    # SEM calculation: stdev / sqrt(n)
    if len(values) > 1:
        sem_val = stats.sem(values)
    else:
        sem_val = 0.0
        
    return {"mean": mean_val, "sem": sem_val, "values": values}

def get_p_value(group_a_vals, group_b_vals):
    if len(group_a_vals) < 2 or len(group_b_vals) < 2:
        return float('nan')
    # Use Student's t-test (equal_var=True)
    t_stat, p_val = stats.ttest_ind(group_a_vals, group_b_vals, equal_var=True)
    return p_val

def format_p_value(p_val):
    if math.isnan(p_val):
        return "N/A"
    
    # Format to 4 decimal places
    if p_val >= 0.0001:
        p_str = f"{p_val:.4f}"
    else:
        p_str = "<0.0001"
        
    # Add significance asterisks
    if p_val < 0.0001:
        p_str += " ****"
    elif p_val < 0.001:
        p_str += " ***"
    elif p_val < 0.01:
        p_str += " **"
    elif p_val < 0.05:
        p_str += " *"
    else:
        p_str += " (ns)"
        
    return p_str

def apply_summary_styles(sheet, start_row, max_col):
    # Styles for table headers and cells
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    double_bottom_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='double', color='000000')
    )
    
    # Header fill and font
    header_fill = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")
    header_font = Font(name="Arial", size=10, bold=True)
    section_font = Font(name="Arial", size=11, bold=True, color="003366")
    
    # Section title style
    sheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=max_col)
    title_cell = sheet.cell(row=start_row, column=1)
    title_cell.font = section_font
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[start_row].height = 25
    
    # Column headers style
    sheet.row_dimensions[start_row + 1].height = 20
    for col in range(1, max_col + 1):
        cell = sheet.cell(row=start_row + 1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

def consolidate():
    cleaned_dir = "/Users/macboook/PROJECTS/Final year project work/bay-leaf-clove-toxicity-study/cleaning/cleaned"
    clove_fp = os.path.join(cleaned_dir, "PHILIP OXIDATIVE STRESS MARKERS RESULTS (2026).xlsx")
    bay_fp = os.path.join(cleaned_dir, "PHILIP OXIDATIVE STRESS MARKERS RESULTS 2 (2026).xlsx")
    
    output_fp = "/Users/macboook/PROJECTS/Final year project work/bay-leaf-clove-toxicity-study/cleaning/OXIDATIVE_STRESS_MARKERS_CONSOLIDATED.xlsx"
    
    # 1. Load Clove workbook
    wb_clove = openpyxl.load_workbook(clove_fp, data_only=True)
    sheet_clove = wb_clove.active
    sheet_clove.title = "Clove"
    
    # 2. Load Bay Leaf workbook and copy to Clove workbook
    wb_bay = openpyxl.load_workbook(bay_fp, data_only=True)
    sheet_bay_src = wb_bay.active
    sheet_bay_dst = wb_clove.create_sheet(title="Bay Leaf")
    
    # Copy all cells and formatting from Bay Leaf source sheet to destination sheet
    for r in range(1, sheet_bay_src.max_row + 1):
        # Set row height
        if r in sheet_bay_src.row_dimensions:
            sheet_bay_dst.row_dimensions[r].height = sheet_bay_src.row_dimensions[r].height
            
        for c in range(1, sheet_bay_src.max_column + 1):
            # Copy column width
            col_letter = openpyxl.utils.get_column_letter(c)
            if col_letter in sheet_bay_src.column_dimensions:
                sheet_bay_dst.column_dimensions[col_letter].width = sheet_bay_src.column_dimensions[col_letter].width
                
            copy_cell(sheet_bay_src.cell(row=r, column=c), sheet_bay_dst.cell(row=r, column=c))
            
    print("Consolidated sheets successfully.")
    
    # 3. Process Clove Sheet Statistics
    # Groups: CL (CL1-15), MC (MC1-5), NC (NC1-5), IB (IB1,3,4,5)
    clove_data = []
    # Read rows
    for r in range(3, 32): # rows 3 to 31 are data rows
        row_vals = [sheet_clove.cell(row=r, column=c).value for c in range(1, 9)]
        clove_data.append(row_vals)
        
    cl_rows = [row for row in clove_data if str(row[0]).startswith("CL")]
    mc_rows = [row for row in clove_data if str(row[0]).startswith("MC")]
    nc_rows = [row for row in clove_data if str(row[0]).startswith("NC")]
    ib_rows = [row for row in clove_data if str(row[0]).startswith("IB")]
    
    print(f"Clove sizes: CL={len(cl_rows)}, MC={len(mc_rows)}, NC={len(nc_rows)}, IB={len(ib_rows)}")
    
    # Compute Clove statistics
    clove_stats = {}
    for col in range(2, 9):
        cl_stat = calculate_group_stats(cl_rows, col)
        mc_stat = calculate_group_stats(mc_rows, col)
        nc_stat = calculate_group_stats(nc_rows, col)
        ib_stat = calculate_group_stats(ib_rows, col)
        
        # Tukey HSD p-values
        # Group order: NC (0), MC (1), CL (2), IB (3)
        res = stats.tukey_hsd(nc_stat["values"], mc_stat["values"], cl_stat["values"], ib_stat["values"])
        p_mc_vs_nc = res.pvalue[1, 0]
        p_cl_vs_mc = res.pvalue[2, 1]
        p_ib_vs_mc = res.pvalue[3, 1]
        p_cl_vs_ib = res.pvalue[2, 3]
        
        clove_stats[col] = {
            "nc": nc_stat,
            "mc": mc_stat,
            "cl": cl_stat,
            "ib": ib_stat,
            "p_mc_nc": p_mc_vs_nc,
            "p_cl_mc": p_cl_vs_mc,
            "p_ib_mc": p_ib_vs_mc,
            "p_cl_ib": p_cl_vs_ib
        }
        
    # Append Clove stats to sheet
    start_row = 33
    sheet_clove.cell(row=start_row, column=1, value="STATISTICAL ANALYSIS SUMMARY")
    
    # Headers
    headers = ["Group / Parameter"] + [sheet_clove.cell(row=1, column=c).value for c in range(2, 9)]
    for idx, h in enumerate(headers, 1):
        sheet_clove.cell(row=start_row + 1, column=idx, value=h)
        
    apply_summary_styles(sheet_clove, start_row, 8)
    
    # Mean +- SEM rows
    groups_to_write = [
        ("NC (Normal Control)", "nc"),
        ("MC (Model Control)", "mc"),
        ("CL (Clove Extract)", "cl"),
        ("IB (Ibuprofen)", "ib")
    ]
    
    border_cell = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    font_bold = Font(name="Arial", size=9, bold=True)
    font_regular = Font(name="Arial", size=9, bold=False)
    
    curr_row = start_row + 2
    for label, key in groups_to_write:
        sheet_clove.cell(row=curr_row, column=1, value=label).font = font_bold
        sheet_clove.cell(row=curr_row, column=1).alignment = Alignment(horizontal="left")
        sheet_clove.cell(row=curr_row, column=1).border = border_cell
        sheet_clove.row_dimensions[curr_row].height = 18
        
        for col in range(2, 9):
            stat = clove_stats[col][key]
            mean_sem_str = f"{stat['mean']:.4f} ± {stat['sem']:.4f}"
            cell = sheet_clove.cell(row=curr_row, column=col, value=mean_sem_str)
            cell.font = font_regular
            cell.alignment = Alignment(horizontal="center")
            cell.border = border_cell
        curr_row += 1
        
    # p-value rows
    p_values_to_write = [
        ("p-value (MC vs NC)", "p_mc_nc"),
        ("p-value (CL vs MC)", "p_cl_mc"),
        ("p-value (IB vs MC)", "p_ib_mc"),
        ("p-value (CL vs IB)", "p_cl_ib")
    ]
    
    for label, key in p_values_to_write:
        sheet_clove.cell(row=curr_row, column=1, value=label).font = font_bold
        sheet_clove.cell(row=curr_row, column=1).alignment = Alignment(horizontal="left")
        sheet_clove.cell(row=curr_row, column=1).border = border_cell
        sheet_clove.row_dimensions[curr_row].height = 18
        
        for col in range(2, 9):
            p_val = clove_stats[col][key]
            p_str = format_p_value(p_val)
            cell = sheet_clove.cell(row=curr_row, column=col, value=p_str)
            cell.font = font_regular
            cell.alignment = Alignment(horizontal="center")
            cell.border = border_cell
        curr_row += 1
        
    print("Clove sheet statistics written successfully.")
    
    # 4. Process Bay Leaf Sheet Statistics
    # Groups: BL (BL1-15), IBU (IBU1,2,4), MG (MG1-3), NG (NG1-3), CL (CL16-17)
    # We parse the groups using the prefix tracker
    bay_data = []
    active_prefix = None
    for r in range(3, 29): # rows 3 to 28 are data rows
        row_id_val = sheet_bay_dst.cell(row=r, column=1).value
        val_str = str(row_id_val).strip()
        has_letters = any(c.isalpha() for c in val_str)
        if has_letters:
            prefix = ""
            for char in val_str:
                if char.isalpha():
                    prefix += char
                elif char.isspace() or char.isdigit():
                    break
            active_prefix = prefix
            
        row_vals = [sheet_bay_dst.cell(row=r, column=c).value for c in range(1, 9)]
        # Add the active prefix to the data list for grouping
        bay_data.append((active_prefix, row_vals))
        
    bl_rows = [row[1] for row in bay_data if row[0] == "BL"]
    ibu_rows = [row[1] for row in bay_data if row[0] == "IBU"]
    mg_rows = [row[1] for row in bay_data if row[0] == "MG"]
    ng_rows = [row[1] for row in bay_data if row[0] == "NG"]
    cl_bay_rows = [row[1] for row in bay_data if row[0] == "CL"]
    
    print(f"Bay Leaf sizes: BL={len(bl_rows)}, IBU={len(ibu_rows)}, MG={len(mg_rows)}, NG={len(ng_rows)}, CL={len(cl_bay_rows)}")
    
    bay_stats = {}
    for col in range(2, 9):
        bl_stat = calculate_group_stats(bl_rows, col)
        ibu_stat = calculate_group_stats(ibu_rows, col)
        mg_stat = calculate_group_stats(mg_rows, col)
        ng_stat = calculate_group_stats(ng_rows, col)
        
        # Tukey HSD p-values
        # Group order: NG (0), MG (1), BL (2), IBU (3)
        res_bay = stats.tukey_hsd(ng_stat["values"], mg_stat["values"], bl_stat["values"], ibu_stat["values"])
        p_mg_vs_ng = res_bay.pvalue[1, 0]
        p_bl_vs_mg = res_bay.pvalue[2, 1]
        p_ibu_vs_mg = res_bay.pvalue[3, 1]
        p_bl_vs_ibu = res_bay.pvalue[2, 3]
        
        bay_stats[col] = {
            "ng": ng_stat,
            "mg": mg_stat,
            "bl": bl_stat,
            "ibu": ibu_stat,
            "p_mg_ng": p_mg_vs_ng,
            "p_bl_mg": p_bl_vs_mg,
            "p_ibu_mg": p_ibu_vs_mg,
            "p_bl_ibu": p_bl_vs_ibu
        }
        
    # Append Bay Leaf stats to sheet
    start_row = 30
    sheet_bay_dst.cell(row=start_row, column=1, value="STATISTICAL ANALYSIS SUMMARY")
    
    # Headers
    headers_bay = ["Group / Parameter"] + [sheet_bay_dst.cell(row=1, column=c).value for c in range(2, 9)]
    for idx, h in enumerate(headers_bay, 1):
        sheet_bay_dst.cell(row=start_row + 1, column=idx, value=h)
        
    apply_summary_styles(sheet_bay_dst, start_row, 8)
    
    # Mean +- SEM rows
    groups_to_write_bay = [
        ("NG (Normal Control)", "ng"),
        ("MG (Model Control)", "mg"),
        ("BL (Bay Leaf Extract)", "bl"),
        ("IBU (Ibuprofen)", "ibu")
    ]
    
    curr_row = start_row + 2
    for label, key in groups_to_write_bay:
        sheet_bay_dst.cell(row=curr_row, column=1, value=label).font = font_bold
        sheet_bay_dst.cell(row=curr_row, column=1).alignment = Alignment(horizontal="left")
        sheet_bay_dst.cell(row=curr_row, column=1).border = border_cell
        sheet_bay_dst.row_dimensions[curr_row].height = 18
        
        for col in range(2, 9):
            stat = bay_stats[col][key]
            mean_sem_str = f"{stat['mean']:.4f} ± {stat['sem']:.4f}"
            cell = sheet_bay_dst.cell(row=curr_row, column=col, value=mean_sem_str)
            cell.font = font_regular
            cell.alignment = Alignment(horizontal="center")
            cell.border = border_cell
        curr_row += 1
        
    # p-value rows
    p_values_to_write_bay = [
        ("p-value (MG vs NG)", "p_mg_ng"),
        ("p-value (BL vs MG)", "p_bl_mg"),
        ("p-value (IBU vs MG)", "p_ibu_mg"),
        ("p-value (BL vs IBU)", "p_bl_ibu")
    ]
    
    for label, key in p_values_to_write_bay:
        sheet_bay_dst.cell(row=curr_row, column=1, value=label).font = font_bold
        sheet_bay_dst.cell(row=curr_row, column=1).alignment = Alignment(horizontal="left")
        sheet_bay_dst.cell(row=curr_row, column=1).border = border_cell
        sheet_bay_dst.row_dimensions[curr_row].height = 18
        
        for col in range(2, 9):
            p_val = bay_stats[col][key]
            p_str = format_p_value(p_val)
            cell = sheet_bay_dst.cell(row=curr_row, column=col, value=p_str)
            cell.font = font_regular
            cell.alignment = Alignment(horizontal="center")
            cell.border = border_cell
        curr_row += 1
        
    print("Bay Leaf sheet statistics written successfully.")
    
    # Save the consolidated workbook
    wb_clove.save(output_fp)
    print(f"Consolidated file saved successfully to: {output_fp}")

if __name__ == "__main__":
    consolidate()
