import openpyxl
import os

def clean_excel_files():
    input_dir = "/Users/macboook/PROJECTS/Final year project work/bay-leaf-clove-toxicity-study/DATA/sheets"
    output_dir = "/Users/macboook/PROJECTS/Final year project work/bay-leaf-clove-toxicity-study/cook/cleaned"
    
    # Create the output directory if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)
    
    for filename in os.listdir(input_dir):
        if filename.endswith(".xlsx") and not filename.startswith("~$"):
            input_path = os.path.join(input_dir, filename)
            output_path = os.path.join(output_dir, filename)
            
            print(f"\nProcessing: {filename}")
            
            # Load workbook with data_only=True to evaluate formulas to values
            wb = openpyxl.load_workbook(input_path, data_only=True)
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                max_col = sheet.max_column
                
                # Determine columns to keep
                columns_to_keep = []
                for col_idx in range(1, max_col + 1):
                    cell = sheet.cell(row=1, column=col_idx)
                    cell_val = cell.value
                    
                    # We keep column 1 ("SAMPLE") or any column where row 1 cell is bold
                    is_bold = (cell.font.bold is True) if cell.font else False
                    if col_idx == 1 or is_bold:
                        # Make sure it actually has a value in row 1
                        if cell_val is not None and str(cell_val).strip() != "":
                            columns_to_keep.append(col_idx)
                
                # Delete columns that are not in columns_to_keep, starting from the rightmost
                deleted_count = 0
                for col_idx in range(max_col, 0, -1):
                    if col_idx not in columns_to_keep:
                        sheet.delete_cols(col_idx)
                        deleted_count += 1
                
                print(f"  Sheet '{sheet_name}': Kept {len(columns_to_keep)} columns (SAMPLE + bolded), stripped {deleted_count} columns.")
            
            # Save the new workbook
            wb.save(output_path)
            print(f"  Saved cleaned file to: {output_path}")

if __name__ == "__main__":
    clean_excel_files()
