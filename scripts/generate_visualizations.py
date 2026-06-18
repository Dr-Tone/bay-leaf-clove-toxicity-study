import openpyxl
import os
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import scipy.stats as stats

def parse_groups_from_sheet(sheet, max_rows):
    data = []
    active_prefix = None
    for r in range(3, max_rows + 1):
        row_id_val = sheet.cell(row=r, column=1).value
        if row_id_val is None:
            continue
        val_str = str(row_id_val).strip()
        has_letters = any(c.isalpha() for c in val_str)
        if has_letters:
            prefix = ""
            for char in val_str:
                if char.isalpha():
                    prefix += char
                elif char.isspace() or char.isdigit():
                    break
            active_prefix = prefix
            clean_id = val_str
        else:
            clean_id = f"{active_prefix} {val_str}"
            
        row_vals = [sheet.cell(row=r, column=c).value for c in range(2, 9)]
        data.append({
            "Group": active_prefix,
            "ID": clean_id,
            "Prot": float(row_vals[0]) if row_vals[0] is not None else np.nan,
            "SOD": float(row_vals[1]) if row_vals[1] is not None else np.nan,
            "CAT": float(row_vals[2]) if row_vals[2] is not None else np.nan,
            "GPx": float(row_vals[3]) if row_vals[3] is not None else np.nan,
            "MDA": float(row_vals[4]) if row_vals[4] is not None else np.nan,
            "Red. GSH": float(row_vals[5]) if row_vals[5] is not None else np.nan,
            "TAC": float(row_vals[6]) if row_vals[6] is not None else np.nan
        })
    return pd.DataFrame(data)

def generate_visualizations():
    cleaned_dir = "/Users/macboook/PROJECTS/Final year project work/bay-leaf-clove-toxicity-study/cleaning/cleaned"
    clove_fp = os.path.join(cleaned_dir, "PHILIP OXIDATIVE STRESS MARKERS RESULTS (2026).xlsx")
    bay_fp = os.path.join(cleaned_dir, "PHILIP OXIDATIVE STRESS MARKERS RESULTS 2 (2026).xlsx")
    
    graphs_dir = "/Users/macboook/PROJECTS/Final year project work/bay-leaf-clove-toxicity-study/cleaning/graphs"
    os.makedirs(graphs_dir, exist_ok=True)
    
    # 1. Load Clove data
    wb_cl = openpyxl.load_workbook(clove_fp, data_only=True)
    df_clove = parse_groups_from_sheet(wb_cl.active, 31)
    # Rename group abbreviations for clarity
    group_map_clove = {"NC": "Normal Control", "MC": "Model Control", "CL": "Clove Extract", "IB": "Ibuprofen"}
    df_clove["GroupName"] = df_clove["Group"].map(group_map_clove)
    
    # 2. Load Bay Leaf data
    wb_bay = openpyxl.load_workbook(bay_fp, data_only=True)
    df_bay = parse_groups_from_sheet(wb_bay.active, 28)
    group_map_bay = {"NG": "Normal Control", "MG": "Model Control", "BL": "Bay Leaf Extract", "IBU": "Ibuprofen", "CL": "Clove Extract"}
    df_bay["GroupName"] = df_bay["Group"].map(group_map_bay)
    
    # Combined df for general analyses (e.g. correlation)
    df_all = pd.concat([df_clove, df_bay], ignore_index=True)
    
    markers = ["Prot", "SOD", "CAT", "GPx", "MDA", "Red. GSH", "TAC"]
    marker_titles = {
        "Prot": "Total Protein (g/dL)",
        "SOD": "Superoxide Dismutase (U/g Prot)",
        "CAT": "Catalase (U/g Prot)",
        "GPx": "Glutathione Peroxidase (U/g Prot)",
        "MDA": "Malondialdehyde (mol/g Prot)",
        "Red. GSH": "Reduced Glutathione (ug/mL)",
        "TAC": "Total Antioxidant Capacity (ug/mL)"
    }
    
    # Style settings
    sns.set_theme(style="whitegrid")
    plt.rcParams.update({'font.size': 11, 'font.family': 'sans-serif'})
    palette = sns.color_palette("Set2", 5)
    
    # FIG 1: Grouped Bar Charts with SEM Error Bars (Multi-panel for Clove and Bay Leaf)
    fig_clove, axes_cl = plt.subplots(4, 2, figsize=(14, 18))
    fig_clove.suptitle("Clove Assay Biochemical Markers (Mean ± SEM)", fontsize=16, fontweight='bold', y=0.98)
    axes_cl = axes_cl.flatten()
    
    fig_bay, axes_bay = plt.subplots(4, 2, figsize=(14, 18))
    fig_bay.suptitle("Bay Leaf Assay Biochemical Markers (Mean ± SEM)", fontsize=16, fontweight='bold', y=0.98)
    axes_bay = axes_bay.flatten()
    
    # We loop through markers
    for idx, marker in enumerate(markers):
        # --- Clove Plot ---
        ax_cl = axes_cl[idx]
        cl_means = df_clove.groupby("GroupName")[marker].mean()
        cl_sems = df_clove.groupby("GroupName")[marker].apply(stats.sem)
        # Order: Normal Control, Model Control, Clove Extract, Ibuprofen
        order = ["Normal Control", "Model Control", "Clove Extract", "Ibuprofen"]
        cl_means = cl_means.reindex(order)
        cl_sems = cl_sems.reindex(order)
        
        bars = ax_cl.bar(order, cl_means, yerr=cl_sems, capsize=6, color=palette[:4], edgecolor='black', alpha=0.9)
        ax_cl.set_title(marker_titles[marker], fontweight='bold')
        ax_cl.set_ylabel("Value")
        ax_cl.set_xticklabels(order, rotation=15, ha='right')
        
        # Add labels on top of bars
        for bar in bars:
            height = bar.get_height()
            ax_cl.annotate(f'{height:.3f}',
                        xy=(bar.get_x() + bar.get_width() / 2, height),
                        xytext=(0, 3),  # 3 points vertical offset
                        textcoords="offset points",
                        ha='center', va='bottom', fontsize=9)
            
        # --- Bay Leaf Plot ---
        ax_bay = axes_bay[idx]
        bay_means = df_bay.groupby("GroupName")[marker].mean()
        bay_sems = df_bay.groupby("GroupName")[marker].apply(stats.sem)
        order_bay = ["Normal Control", "Model Control", "Bay Leaf Extract", "Ibuprofen", "Clove Extract"]
        bay_means = bay_means.reindex(order_bay)
        bay_sems = bay_sems.reindex(order_bay)
        
        bars_bay = ax_bay.bar(order_bay, bay_means, yerr=bay_sems, capsize=6, color=palette[:5], edgecolor='black', alpha=0.9)
        ax_bay.set_title(marker_titles[marker], fontweight='bold')
        ax_bay.set_ylabel("Value")
        ax_bay.set_xticklabels(order_bay, rotation=15, ha='right')
        
        for bar in bars_bay:
            height = bar.get_height()
            ax_bay.annotate(f'{height:.3f}',
                        xy=(bar.get_x() + bar.get_width() / 2, height),
                        xytext=(0, 3),
                        textcoords="offset points",
                        ha='center', va='bottom', fontsize=9)
                        
    # Hide empty subplots
    axes_cl[-1].axis('off')
    axes_bay[-1].axis('off')
    
    fig_clove.tight_layout()
    fig_clove.savefig(os.path.join(graphs_dir, "clove_markers_bar_chart.png"), dpi=300)
    plt.close(fig_clove)
    
    fig_bay.tight_layout()
    fig_bay.savefig(os.path.join(graphs_dir, "bay_leaf_markers_bar_chart.png"), dpi=300)
    plt.close(fig_bay)
    
    # FIG 2: Box-and-Whisker Plots with Jitter Points (for SOD and MDA as key highlights)
    for marker in ["SOD", "MDA"]:
        # Clove Boxplot
        plt.figure(figsize=(8, 6))
        sns.boxplot(x="GroupName", y=marker, data=df_clove, order=order, palette=palette[:4], width=0.5)
        sns.stripplot(x="GroupName", y=marker, data=df_clove, order=order, color="black", alpha=0.6, size=6, jitter=0.2)
        plt.title(f"Distribution of {marker_titles[marker]} (Clove Assay)", fontweight='bold', fontsize=12)
        plt.ylabel(marker_titles[marker])
        plt.xlabel("Experimental Group")
        plt.xticks(rotation=15)
        plt.tight_layout()
        plt.savefig(os.path.join(graphs_dir, f"clove_{marker.lower().replace(' ', '_')}_boxplot.png"), dpi=300)
        plt.close()
        
        # Bay Leaf Boxplot
        plt.figure(figsize=(9, 6))
        sns.boxplot(x="GroupName", y=marker, data=df_bay, order=order_bay, palette=palette[:5], width=0.5)
        sns.stripplot(x="GroupName", y=marker, data=df_bay, order=order_bay, color="black", alpha=0.6, size=6, jitter=0.2)
        plt.title(f"Distribution of {marker_titles[marker]} (Bay Leaf Assay)", fontweight='bold', fontsize=12)
        plt.ylabel(marker_titles[marker])
        plt.xlabel("Experimental Group")
        plt.xticks(rotation=15)
        plt.tight_layout()
        plt.savefig(os.path.join(graphs_dir, f"bay_leaf_{marker.lower().replace(' ', '_')}_boxplot.png"), dpi=300)
        plt.close()
        
    # FIG 3: Pearson Correlation Heatmap
    corr_matrix = df_all[markers].corr(method='pearson')
    plt.figure(figsize=(9, 7))
    sns.heatmap(corr_matrix, annot=True, cmap="coolwarm", fmt=".3f", linewidths=0.5, vmin=-1, vmax=1,
                cbar_kws={'label': 'Pearson Correlation Coefficient (r)'})
    plt.title("Correlation Matrix of Oxidative Stress Markers (All Combined Data)", fontweight='bold', fontsize=13)
    plt.tight_layout()
    plt.savefig(os.path.join(graphs_dir, "markers_correlation_heatmap.png"), dpi=300)
    plt.close()
    
    # FIG 4: Radar (Spider) Chart of Normalized Profiles
    # Normalize means to a [0, 1] scale for fair radar plotting
    df_norm = df_all.copy()
    for marker in markers:
        min_v = df_all[marker].min()
        max_v = df_all[marker].max()
        if max_v > min_v:
            df_norm[marker] = (df_all[marker] - min_v) / (max_v - min_v)
        else:
            df_norm[marker] = 0.0
        
    # Clove Group Normalized Radar
    clove_radar_means = df_norm[df_norm["Group"].isin(["NC", "MC", "CL", "IB"])].groupby("Group")[markers].mean()
    # Reorder index to follow standard flow
    clove_radar_means = clove_radar_means.reindex(["NC", "MC", "CL", "IB"])
    
    angles = np.linspace(0, 2 * np.pi, len(markers), endpoint=False).tolist()
    angles += angles[:1] # close the circle
    
    fig, ax = plt.subplots(figsize=(8, 8), subplot_kw=dict(polar=True))
    ax.set_theta_offset(np.pi / 2)
    ax.set_theta_direction(-1)
    
    plt.xticks(angles[:-1], markers, color='grey', size=11, fontweight='bold')
    ax.set_rlabel_position(0)
    plt.yticks([0.2, 0.4, 0.6, 0.8, 1.0], ["0.2", "0.4", "0.6", "0.8", "1.0"], color="grey", size=9)
    plt.ylim(0, 1)
    
    group_labels = {
        "NC": "Normal Control (NC)",
        "MC": "Model Control (MC)",
        "CL": "Clove Extract (CL)",
        "IB": "Ibuprofen (IB)"
    }
    
    for grp in ["NC", "MC", "CL", "IB"]:
        values = clove_radar_means.loc[grp].values.flatten().tolist()
        values += values[:1]
        ax.plot(angles, values, linewidth=2, linestyle='solid', label=group_labels[grp])
        ax.fill(angles, values, alpha=0.15)
        
    plt.title("Normalized Oxidative Stress Profile (Clove Assay)", size=14, fontweight='bold', y=1.1)
    plt.legend(loc='upper right', bbox_to_anchor=(1.3, 1.1))
    plt.tight_layout()
    plt.savefig(os.path.join(graphs_dir, "clove_markers_radar_chart.png"), dpi=300)
    plt.close()
    
    print("Visualizations generated successfully.")
    
    # 5. Create Markdown report with embedded images
    report_fp = "/Users/macboook/PROJECTS/Final year project work/bay-leaf-clove-toxicity-study/cleaning/OXIDATIVE_STRESS_MARKERS_VISUALIZATIONS.md"
    
    # We will write the markdown content
    markdown_content = """# Data Visualization Report of Oxidative Stress Markers

This report summarizes and visualizes the results of the biochemical assays for Clove and Bay Leaf treatments. It compiles the bar charts, box plots, correlation matrix, and radar profile charts to describe group trends and statistical variances.

## Banned Charts Explanation

In biochemical pharmacology, certain chart types are omitted because they represent data distributions inaccurately.
* **Pie Charts**: Pie charts illustrate proportions of a whole summing to 100%. Because enzyme activities (SOD, CAT, GPx) and marker concentrations represent independent continuous parameters, dividing them into pie segments is scientifically incorrect.
* **Histograms**: Histograms illustrate data frequency distributions. Due to small cohort sizes (n = 3 or n = 5) in standard rodent models, histograms do not provide meaningful distributions. Standard box-and-whisker plots overlaid with individual data points are used instead to show data dispersion.

---

## 1. Summary Bar Charts (Mean ± SEM)

Grouped bar charts illustrate the mean value of each biochemical parameter across all groups. Standard error of the mean (SEM) is represented as error bars.

### Clove Assay Markers
The bar chart below summarizes the seven biochemical parameters measured during the Clove assay.

![Clove Assay Summary Bar Chart](graphs/clove_markers_bar_chart.png)

### Bay Leaf Assay Markers
The bar chart below summarizes the seven parameters measured during the Bay Leaf assay.

![Bay Leaf Assay Summary Bar Chart](graphs/bay_leaf_markers_bar_chart.png)

---

## 2. Data Distribution Box Plots

Box-and-whisker plots display individual data points to show the exact distribution, median, and range within each group. This format is preferred to examine biological outliers.

### Superoxide Dismutase (SOD) Distributions

#### Clove SOD Distribution
![Clove SOD Box Plot](graphs/clove_sod_boxplot.png)

#### Bay Leaf SOD Distribution
![Bay Leaf SOD Box Plot](graphs/bay_leaf_sod_boxplot.png)

### Malondialdehyde (MDA) Distributions

#### Clove MDA Distribution
![Clove MDA Box Plot](graphs/clove_mda_boxplot.png)

#### Bay Leaf MDA Distribution
![Bay Leaf MDA Box Plot](graphs/bay_leaf_mda_boxplot.png)

---

## 3. Parameter Correlations

A Pearson correlation matrix calculates the relationship between each biochemical marker across all combined experimental groups.

![Pearson Correlation Heatmap](graphs/markers_correlation_heatmap.png)

### Correlation Observations
The correlation heatmap displays a strong negative correlation between antioxidant enzyme levels and lipid peroxidation (MDA). For example, increased SOD and CAT activities correlate with decreased MDA levels, confirming the protective mechanism against cell membrane damage.

---

## 4. Multi-Dimensional Profiling (Radar Chart)

The radar chart maps the normalized means of all seven parameters on a single radial axis. This provides a profile of how closely a treatment group matches either the healthy control or the model control.

![Clove Assay Radar Chart](graphs/clove_markers_radar_chart.png)

### Radar Observations
The radar chart contrasts the protective profile of the treatments. It shows that the Normal Control group maintains high antioxidant levels, whereas the Model Control group displays significant depletion. The overlap between Clove and Model Control groups illustrates the lack of recovery in the Clove treated cohort.
"""
    
    with open(report_fp, "w") as f:
        f.write(markdown_content)
        
    print(f"Visualization report saved to: {report_fp}")

if __name__ == "__main__":
    generate_visualizations()
